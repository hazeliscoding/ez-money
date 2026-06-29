"""Render cleaned transactions into the formatted 2026 tracker workbook.

Pure output layer: give it a list[Transaction] and it writes the 5-tab,
formula-driven .xlsx (Dashboard / Transactions / Monthly Summary / Setup /
Guide). Supports any number of statement periods in one workbook.
"""

import datetime as dt

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import DataBarRule, CellIsRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

from .config import CATEGORIES, DEFAULT_BUDGETS, INCOME_LABEL

# ---- palette ----
NAVY, BLUE, LBLUE, LBLUE2 = "1F3864", "2F5496", "D9E1F2", "EAF0FA"
GOLD, GREEN, LGREEN, RED, LRED = "BF9000", "548235", "C6EFCE", "C00000", "FFC7CE"
GREY, WHITE, CARD_BG = "808080", "FFFFFF", "F2F6FC"

MONEY = '_($* #,##0.00_);[Red]_($* (#,##0.00);_($* "-"??_);_(@_)'
MONEY0 = '$#,##0'
PCT = '0.0%'
DATEFMT = 'mm/dd/yyyy'
MONTHS3 = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _fill(h):
    return PatternFill("solid", fgColor=h)


def _font(**k):
    return Font(name="Calibri", **k)


_thin = Side(style="thin", color="BFBFBF")
_med = Side(style="medium", color=BLUE)
_box = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def read_existing_budgets(path: str) -> dict | None:
    """Read user-edited budgets from an existing workbook so re-imports keep them."""
    try:
        wb = load_workbook(path, data_only=True)
        s = wb["Setup"]
    except Exception:
        return None
    out = {}
    for r in range(6, 6 + len(CATEGORIES) + 1):
        name = s.cell(r, 2).value
        val = s.cell(r, 3).value
        if name in CATEGORIES and isinstance(val, (int, float)):
            out[name] = val
    return out or None


def _infer(transactions, year, default_period):
    if not transactions:
        y = year or 2026
        return y, (default_period or f"Jan {y}")
    latest = max(transactions, key=lambda t: t.date)
    if year is None:
        year = int(latest.period.split()[-1])
    if default_period is None:
        default_period = latest.period
    return year, default_period


def build_workbook(transactions, out_path, budgets=None,
                   default_period=None, year=None):
    budgets = {**DEFAULT_BUDGETS, **(budgets or {})}
    year, default_period = _infer(transactions, year, default_period)
    periods = [f"{m} {year}" for m in MONTHS3]
    txns = sorted(transactions, key=lambda t: (t.date, t.description))

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_tx = wb.create_sheet("Transactions")
    ws_sum = wb.create_sheet("Monthly Summary")
    ws_set = wb.create_sheet("Setup")
    ws_help = wb.create_sheet("Guide")
    for ws in (ws_dash, ws_tx, ws_sum, ws_set, ws_help):
        ws.sheet_view.showGridLines = False

    cl = get_column_letter

    # ============================== SETUP ==============================
    s = ws_set
    s["B2"] = "Setup & Reference"
    s["B2"].font = _font(size=16, bold=True, color=NAVY)
    s["B3"] = "Edit budgets here. These categories drive the dropdowns and every summary."
    s["B3"].font = _font(size=10, italic=True, color=GREY)
    s["B5"], s["C5"] = "Category", "Monthly Budget"
    for c in ("B5", "C5"):
        s[c].font = _font(bold=True, color=WHITE)
        s[c].fill = _fill(BLUE)
        s[c].alignment = Alignment(horizontal="center")
        s[c].border = _box
    r0 = 6
    for i, cat in enumerate(CATEGORIES):
        rr = r0 + i
        band = LBLUE2 if i % 2 else WHITE
        s.cell(rr, 2, cat).border = _box
        s.cell(rr, 2).fill = _fill(band)
        bc = s.cell(rr, 3, budgets.get(cat, 0))
        bc.number_format = MONEY
        bc.border = _box
        bc.fill = _fill(band)
    cat_last = r0 + len(CATEGORIES) - 1
    s.cell(cat_last + 1, 2, "Total monthly budget").font = _font(bold=True)
    tb = s.cell(cat_last + 1, 3, f"=SUM(C{r0}:C{cat_last})")
    tb.number_format = MONEY
    tb.font = _font(bold=True, color=NAVY)
    tb.fill = _fill(LBLUE)
    s.cell(cat_last + 1, 2).fill = _fill(LBLUE)
    CAT_RANGE = f"Setup!$B${r0}:$B${cat_last}"
    BUD_RANGE = f"Setup!$B${r0}:$C${cat_last}"
    TOTBUD_CELL = f"Setup!$C${cat_last + 1}"

    s["E5"] = "Type"
    s["E5"].font = _font(bold=True, color=WHITE)
    s["E5"].fill = _fill(BLUE)
    s["E5"].alignment = Alignment(horizontal="center")
    s["E5"].border = _box
    s["E6"], s["E7"] = "Expense", "Income"
    for rr in (6, 7):
        s.cell(rr, 5).border = _box
    TYPE_RANGE = "Setup!$E$6:$E$7"

    s["G5"] = "Statement Periods"
    s["G5"].font = _font(bold=True, color=WHITE)
    s["G5"].fill = _fill(BLUE)
    s["G5"].alignment = Alignment(horizontal="center")
    s["G5"].border = _box
    for i, p in enumerate(periods):
        pc = s.cell(6 + i, 7, p)
        pc.border = _box
        pc.alignment = Alignment(horizontal="center")
    PERIOD_RANGE = f"Setup!$G$6:$G${6 + len(periods) - 1}"

    s["I5"] = "Category pick-list"
    s["I5"].font = _font(bold=True, color=WHITE)
    s["I5"].fill = _fill(BLUE)
    s["I5"].alignment = Alignment(horizontal="center")
    s["I5"].border = _box
    picks = CATEGORIES + [INCOME_LABEL]
    for i, p in enumerate(picks):
        s.cell(6 + i, 9, p).border = _box
    PICK_RANGE = f"Setup!$I$6:$I${6 + len(picks) - 1}"

    for col, w in {"A": 2, "B": 24, "C": 16, "D": 3, "E": 12,
                   "F": 3, "G": 18, "H": 3, "I": 22}.items():
        s.column_dimensions[col].width = w

    # ============================ TRANSACTIONS =========================
    t = ws_tx
    t["A1"] = "Transactions"
    t["A1"].font = _font(size=16, bold=True, color=NAVY)
    t["F1"] = ("Add new rows at the bottom — summaries update automatically. "
               "Amounts are positive; Type sets income vs. spending.")
    t["F1"].font = _font(size=10, italic=True, color=GREY)
    headers = ["Date", "Period", "Description", "Category", "Type",
               "Amount", "Account", "Notes"]
    hr = 3
    for j, h in enumerate(headers, 1):
        c = t.cell(hr, j, h)
        c.font = _font(bold=True, color=WHITE)
        c.fill = _fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
    fd = hr + 1
    for i, tx in enumerate(txns):
        rr = fd + i
        t.cell(rr, 1, tx.date).number_format = DATEFMT
        t.cell(rr, 2, tx.period)
        t.cell(rr, 3, tx.description)
        t.cell(rr, 4, tx.category)
        t.cell(rr, 5, tx.kind)
        t.cell(rr, 6, tx.amount).number_format = MONEY
        t.cell(rr, 7, tx.account)
        t.cell(rr, 8, tx.notes)
        t.cell(rr, 1).alignment = Alignment(horizontal="center")
        t.cell(rr, 2).alignment = Alignment(horizontal="center")
        t.cell(rr, 5).alignment = Alignment(horizontal="center")
    ld = fd + len(txns) - 1 if txns else fd

    table = Table(displayName="Transactions", ref=f"A{hr}:H{ld}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False,
        showFirstColumn=False, showLastColumn=False)
    t.add_table(table)
    if txns:
        t.conditional_formatting.add(
            f"E{fd}:E{ld}",
            CellIsRule(operator="equal", formula=['"Income"'],
                       fill=_fill(LGREEN), font=_font(color=GREEN, bold=True)))

    dv_cat = DataValidation(type="list", formula1=f"={PICK_RANGE}", allow_blank=True)
    dv_cat.errorStyle = "warning"
    dv_typ = DataValidation(type="list", formula1=f"={TYPE_RANGE}", allow_blank=True)
    dv_per = DataValidation(type="list", formula1=f"={PERIOD_RANGE}", allow_blank=True)
    for dv in (dv_cat, dv_typ, dv_per):
        t.add_data_validation(dv)
    dv_per.add(f"B{fd}:B2000")
    dv_cat.add(f"D{fd}:D2000")
    dv_typ.add(f"E{fd}:E2000")
    t.freeze_panes = "A4"
    for col, w in {"A": 12, "B": 11, "C": 30, "D": 22, "E": 11,
                   "F": 13, "G": 16, "H": 24}.items():
        t.column_dimensions[col].width = w

    # ========================== MONTHLY SUMMARY ========================
    m = ws_sum
    m["A1"] = "Monthly Summary — " + str(year)
    m["A1"].font = _font(size=16, bold=True, color=NAVY)
    m["A2"] = ("Live formulas over the Transactions table. New statement periods "
               "fill in as you add data.")
    m["A2"].font = _font(size=10, italic=True, color=GREY)
    hrow = 4
    hc = m.cell(hrow, 1, "Category")
    hc.font = _font(bold=True, color=WHITE)
    hc.fill = _fill(NAVY)
    hc.alignment = Alignment(horizontal="left", vertical="center")
    for mi in range(12):
        c = m.cell(hrow, 2 + mi, f"=Setup!$G${6 + mi}")
        c.font = _font(bold=True, color=WHITE)
        c.fill = _fill(NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ytd = 14
    yc = m.cell(hrow, ytd, "YTD Total")
    yc.font = _font(bold=True, color=WHITE)
    yc.fill = _fill(GOLD)
    yc.alignment = Alignment(horizontal="center", vertical="center")

    cr0 = hrow + 1
    for i, cat in enumerate(CATEGORIES):
        rr = cr0 + i
        band = LBLUE2 if i % 2 else WHITE
        cc = m.cell(rr, 1, cat)
        cc.fill = _fill(band)
        cc.border = _box
        for mi in range(12):
            col = 2 + mi
            ml = cl(col)
            c = m.cell(rr, col,
                       f'=SUMIFS(Transactions[Amount],Transactions[Type],"Expense",'
                       f'Transactions[Category],$A{rr},Transactions[Period],{ml}${hrow})')
            c.number_format = MONEY
            c.fill = _fill(band)
            c.border = _box
        ycc = m.cell(rr, ytd, f"=SUM(B{rr}:M{rr})")
        ycc.number_format = MONEY
        ycc.font = _font(bold=True)
        ycc.fill = _fill("FFF2CC")
        ycc.border = _box
    clast = cr0 + len(CATEGORIES) - 1

    te = clast + 1
    m.cell(te, 1, "Total Expenses").font = _font(bold=True, color=WHITE)
    m.cell(te, 1).fill = _fill(BLUE)
    for mi in range(12):
        col = 2 + mi
        ml = cl(col)
        c = m.cell(te, col, f"=SUM({ml}{cr0}:{ml}{clast})")
        c.number_format = MONEY
        c.font = _font(bold=True, color=WHITE)
        c.fill = _fill(BLUE)
    c = m.cell(te, ytd, f"=SUM(B{te}:M{te})")
    c.number_format = MONEY
    c.font = _font(bold=True, color=WHITE)
    c.fill = _fill(BLUE)

    ti = te + 1
    m.cell(ti, 1, "Total Income").font = _font(bold=True, color=WHITE)
    m.cell(ti, 1).fill = _fill(GREEN)
    for mi in range(12):
        col = 2 + mi
        ml = cl(col)
        c = m.cell(ti, col, f'=SUMIFS(Transactions[Amount],Transactions[Type],"Income",'
                            f'Transactions[Period],{ml}${hrow})')
        c.number_format = MONEY
        c.font = _font(bold=True, color=WHITE)
        c.fill = _fill(GREEN)
    c = m.cell(ti, ytd, f"=SUM(B{ti}:M{ti})")
    c.number_format = MONEY
    c.font = _font(bold=True, color=WHITE)
    c.fill = _fill(GREEN)

    nr = ti + 1
    m.cell(nr, 1, "Net (Income − Expenses)").font = _font(bold=True, color=NAVY)
    m.cell(nr, 1).fill = _fill(LBLUE)
    for mi in range(12):
        col = 2 + mi
        ml = cl(col)
        c = m.cell(nr, col, f"={ml}{ti}-{ml}{te}")
        c.number_format = MONEY
        c.font = _font(bold=True)
        c.fill = _fill(LBLUE)
    c = m.cell(nr, ytd, f"={cl(ytd)}{ti}-{cl(ytd)}{te}")
    c.number_format = MONEY
    c.font = _font(bold=True)
    c.fill = _fill(LBLUE)

    srr = nr + 1
    m.cell(srr, 1, "Savings Rate").font = _font(italic=True, color=GREY)
    for mi in range(12):
        col = 2 + mi
        ml = cl(col)
        c = m.cell(srr, col, f'=IFERROR({ml}{nr}/{ml}{ti},"")')
        c.number_format = PCT
        c.font = _font(italic=True, color=GREY)
    c = m.cell(srr, ytd, f'=IFERROR({cl(ytd)}{nr}/{cl(ytd)}{ti},"")')
    c.number_format = PCT
    c.font = _font(italic=True, color=GREY)

    m.conditional_formatting.add(
        f"B{nr}:N{nr}",
        CellIsRule(operator="lessThan", formula=["0"], font=_font(bold=True, color=RED)))
    m.freeze_panes = "B5"
    m.column_dimensions["A"].width = 24
    for mi in range(12):
        m.column_dimensions[cl(2 + mi)].width = 12
    m.column_dimensions[cl(ytd)].width = 14

    lc = LineChart()
    lc.title = "Income vs Expenses by Period"
    lc.style = 12
    lc.height = 7.5
    lc.width = 24
    lc.y_axis.numFmt = '$#,##0'
    cats = Reference(m, min_col=2, max_col=13, min_row=hrow, max_row=hrow)
    lc.add_data(Reference(m, min_col=1, max_col=13, min_row=te, max_row=te),
                titles_from_data=True, from_rows=True)
    lc.add_data(Reference(m, min_col=1, max_col=13, min_row=ti, max_row=ti),
                titles_from_data=True, from_rows=True)
    lc.set_categories(cats)
    m.add_chart(lc, f"A{srr + 3}")

    # ============================== DASHBOARD ==========================
    d = ws_dash
    d.column_dimensions["A"].width = 2
    for col in ("B", "C", "D", "E", "F", "G"):
        d.column_dimensions[col].width = 17
    d.column_dimensions["B"].width = 22

    d.merge_cells("B2:G2")
    d["B2"] = str(year) + " Financial Tracker"
    d["B2"].font = _font(size=22, bold=True, color=WHITE)
    d["B2"].alignment = Alignment(horizontal="left", vertical="center")
    d.merge_cells("B3:G3")
    d["B3"] = "Budget & Spending Overview"
    d["B3"].font = _font(size=11, color=WHITE)
    d["B3"].alignment = Alignment(horizontal="left", vertical="center")
    for r in (2, 3):
        for col in range(2, 8):
            d.cell(r, col).fill = _fill(NAVY)
    d.row_dimensions[2].height = 32
    d.row_dimensions[3].height = 18

    d["B5"] = "Viewing period:"
    d["B5"].font = _font(bold=True, size=11, color=NAVY)
    sel = d["C5"]
    sel.value = default_period
    sel.font = _font(bold=True, size=12, color=WHITE)
    sel.fill = _fill(GOLD)
    sel.alignment = Alignment(horizontal="center")
    sel.border = Border(left=_med, right=_med, top=_med, bottom=_med)
    dv_month = DataValidation(type="list", formula1=f"={PERIOD_RANGE}", allow_blank=False)
    d.add_data_validation(dv_month)
    dv_month.add("C5")
    d["D5"] = "◀ pick any period"
    d["D5"].font = _font(italic=True, color=GREY, size=10)
    SEL = "$C$5"

    def kpi(c0, label, formula, accent):
        a, b = cl(c0), cl(c0 + 1)
        d.merge_cells(f"{a}7:{b}7")
        lab = d[f"{a}7"]
        lab.value = label
        lab.font = _font(bold=True, size=10, color=WHITE)
        lab.alignment = Alignment(horizontal="center", vertical="center")
        d.merge_cells(f"{a}8:{b}8")
        val = d[f"{a}8"]
        val.value = formula
        val.number_format = MONEY0
        val.font = _font(bold=True, size=18, color=accent)
        val.alignment = Alignment(horizontal="center", vertical="center")
        for rr in (7, 8):
            for cc in (c0, c0 + 1):
                cell = d.cell(rr, cc)
                cell.fill = _fill(accent) if rr == 7 else _fill(CARD_BG)
                if rr == 8:
                    cell.border = Border(left=_thin, right=_thin, bottom=_thin)
        d.row_dimensions[7].height = 18
        d.row_dimensions[8].height = 30

    inc_f = (f'=SUMIFS(Transactions[Amount],Transactions[Type],"Income",'
             f'Transactions[Period],{SEL})')
    exp_f = (f'=SUMIFS(Transactions[Amount],Transactions[Type],"Expense",'
             f'Transactions[Period],{SEL})')
    kpi(2, "Income (period)", inc_f, GREEN)
    kpi(4, "Spending (period)", exp_f, RED)
    kpi(6, "Net (period)", "=B8-D8", NAVY)

    d["B10"] = "Savings rate:"
    d["B10"].font = _font(bold=True, color=NAVY)
    d["C10"] = '=IFERROR((B8-D8)/B8,"—")'
    d["C10"].number_format = PCT
    d["C10"].font = _font(bold=True, color=GREEN)
    d["E10"] = "Total budget:"
    d["E10"].font = _font(bold=True, color=NAVY)
    d["F10"] = f"={TOTBUD_CELL}"
    d["F10"].number_format = MONEY0
    d["F10"].font = _font(bold=True, color=NAVY)

    bv = 12
    d.merge_cells(f"B{bv}:G{bv}")
    d[f"B{bv}"] = "Budget vs. Actual — selected period"
    d[f"B{bv}"].font = _font(bold=True, size=13, color=NAVY)
    hh = bv + 1
    for j, h in enumerate(["Category", "Budget", "Actual", "Remaining",
                           "% of Budget", "% of Spend"]):
        c = d.cell(hh, 2 + j, h)
        c.font = _font(bold=True, color=WHITE)
        c.fill = _fill(BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _box
    br0 = hh + 1
    for i, cat in enumerate(CATEGORIES):
        rr = br0 + i
        band = LBLUE2 if i % 2 else WHITE
        cc = d.cell(rr, 2, cat)
        cc.border = _box
        cc.fill = _fill(band)
        bud = d.cell(rr, 3, f'=VLOOKUP($B{rr},{BUD_RANGE},2,FALSE)')
        bud.number_format = MONEY
        bud.border = _box
        bud.fill = _fill(band)
        act = d.cell(rr, 4,
                     f'=SUMIFS(Transactions[Amount],Transactions[Type],"Expense",'
                     f'Transactions[Category],$B{rr},Transactions[Period],{SEL})')
        act.number_format = MONEY
        act.border = _box
        act.fill = _fill(band)
        rem = d.cell(rr, 5, f"=C{rr}-D{rr}")
        rem.number_format = MONEY
        rem.border = _box
        rem.fill = _fill(band)
        pob = d.cell(rr, 6, f'=IFERROR(D{rr}/C{rr},"")')
        pob.number_format = PCT
        pob.border = _box
        pob.fill = _fill(band)
        pos = d.cell(rr, 7, f'=IFERROR(D{rr}/$D${br0 + len(CATEGORIES)},"")')
        pos.number_format = PCT
        pos.border = _box
        pos.fill = _fill(band)
    bclast = br0 + len(CATEGORIES) - 1
    tr = bclast + 1
    d.cell(tr, 2, "TOTAL").font = _font(bold=True, color=WHITE)
    d.cell(tr, 2).fill = _fill(NAVY)
    d.cell(tr, 2).border = _box
    for col, formula in ((3, f"=SUM(C{br0}:C{bclast})"),
                         (4, f"=SUM(D{br0}:D{bclast})"),
                         (5, f"=SUM(E{br0}:E{bclast})"),
                         (6, f'=IFERROR(D{tr}/C{tr},"")')):
        c = d.cell(tr, col, formula)
        c.number_format = PCT if col == 6 else MONEY
        c.font = _font(bold=True, color=WHITE)
        c.fill = _fill(NAVY)
        c.border = _box
    d.cell(tr, 7).fill = _fill(NAVY)
    d.cell(tr, 7).border = _box

    d.conditional_formatting.add(
        f"D{br0}:D{bclast}",
        DataBarRule(start_type="num", start_value=0, end_type="max",
                    color="5B9BD5", showValue=True))
    d.conditional_formatting.add(
        f"E{br0}:E{bclast}",
        CellIsRule(operator="lessThan", formula=["0"], fill=_fill(LRED),
                   font=_font(color=RED, bold=True)))
    d.conditional_formatting.add(
        f"F{br0}:F{bclast}",
        CellIsRule(operator="greaterThan", formula=["1"], fill=_fill(LRED),
                   font=_font(color=RED, bold=True)))

    pie = PieChart()
    pie.title = "Where the money goes (selected period)"
    pie.height = 8.5
    pie.width = 11.5
    pie.add_data(Reference(d, min_col=4, min_row=hh, max_row=bclast),
                 titles_from_data=True)
    pie.set_categories(Reference(d, min_col=2, min_row=br0, max_row=bclast))
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    d.add_chart(pie, f"I{hh}")

    bar = BarChart()
    bar.type = "bar"
    bar.title = "Budget vs Actual by Category"
    bar.height = 11
    bar.width = 13
    bar.y_axis.numFmt = '$#,##0'
    bar.add_data(Reference(d, min_col=3, min_row=hh, max_row=bclast), titles_from_data=True)
    bar.add_data(Reference(d, min_col=4, min_row=hh, max_row=bclast), titles_from_data=True)
    bar.set_categories(Reference(d, min_col=2, min_row=br0, max_row=bclast))
    bar.gapWidth = 60
    d.add_chart(bar, f"I{hh + 18}")

    # ================================ GUIDE ============================
    g = ws_help
    g.column_dimensions["A"].width = 2
    g.column_dimensions["B"].width = 104
    g["B2"] = "How to use this workbook"
    g["B2"].font = _font(size=18, bold=True, color=NAVY)
    lines = [
        ("h", "The 5 tabs"),
        ("t", "• Dashboard — at-a-glance view. Pick a period in the gold cell (C5); every number + chart updates."),
        ("t", "• Transactions — the data table. Re-imports rebuild it from your statement PDFs."),
        ("t", "• Monthly Summary — auto grid of category × statement period + an income/expense trend."),
        ("t", "• Setup — categories and editable monthly budgets (preserved across re-imports)."),
        ("t", "• Guide — this page."),
        ("s", ""),
        ("h", "How it's generated"),
        ("t", "Run:  python -m ezmoney import \"docs/statements/<year>/*.pdf\""),
        ("t", "The importer parses each Chime statement, strips internal Credit Builder plumbing"),
        ("t", "(Moved to/from, Round Ups, card auto-payment, payday-advance washes), categorizes the rest,"),
        ("t", "and rebuilds this workbook. Statements are bucketed by PERIOD (the ~29th–28th cycle),"),
        ("t", "so rent and the late-month paycheck stay with the statement they belong to."),
        ("s", ""),
        ("h", "Changing categories so it sticks"),
        ("t", "Edit ezmoney/category_rules.json (ordered [pattern, category] list; first match wins) and"),
        ("t", "re-import. To hide a line entirely, add a substring to the 'exclude' list there."),
        ("t", "Editing a Category cell directly works too, but a re-import would overwrite it."),
    ]
    rr = 4
    for kind, text in lines:
        c = g.cell(rr, 2, text)
        if kind == "h":
            c.font = _font(size=12, bold=True, color=BLUE)
        elif kind == "t":
            c.font = _font(size=10, color="333333")
        rr += 1

    wb.save(out_path)
    return out_path
